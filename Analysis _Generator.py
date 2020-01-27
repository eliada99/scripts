import openpyxl
import os
import urllib
import re
import easygui as eg
import sys
from openpyxl import load_workbook

'''----------------Inputs----------------------------------'''
def inputValidation():
       msg = "Fill the following fields:"
       title = "Analysis generator"
       fieldNames = ["Excel file path:","Test folder path:"]
       fieldValues = []  # we start with blanks for the values
       fieldValues = eg.multenterbox(msg,title, fieldNames)

       # make sure that none of the fields was left blank
       while 1:
              if fieldValues == None:
                     break
              errmsg = ""
              for i in range(len(fieldNames)):
                     if fieldValues[i].strip() == "":
                            errmsg = errmsg + ('"%s" is a required field.\n\n' % fieldNames[i])
              if ".xlsx" not in fieldValues[0]:
                     errmsg = errmsg + ('"%s" please provide full file path.\n\n' % fieldNames[0])
              if errmsg == "":
                     break # no problems found
              fieldValues = eg.multenterbox(errmsg, title, fieldNames, fieldValues)
       return fieldValues
'''--------------Description-------------------------------------------'''
def descriptionInitialize(fileName):
        descriptionContent = 'Test Purpose:'+ "\n"
        descriptionContent = descriptionContent + 'The purpose of the test is to verify the following functionality of ' + fileName + ' class:'+ "\n"
        descriptionContent = descriptionContent + '1.'+ "\n"+ "\n"
        descriptionContent = descriptionContent + 'Test Description:'+ "\n"
        descriptionContent = descriptionContent + 'The test will verify the ' + fileName + ' object functionalities using code inspection.' + "\n"+ "\n"
        descriptionContent = descriptionContent + 'Test Inputs:'+ "\n"
        descriptionContent = descriptionContent + 'Use ' + fileName + ' files for the code inspection.'+ "\n" + "\n" 
        descriptionContent = descriptionContent + 'Conditions (includes initialization):' + "\n"
        descriptionContent = descriptionContent + 'The code is from relevant label.' + "\n"+ "\n"
        descriptionContent = descriptionContent + 'Expected Results and Pass/Fail Criteria:'+ "\n"
        descriptionContent = descriptionContent + 'The DMAP use containers accordingly.' + "\n"+ "\n"
        descriptionContent = descriptionContent + 'Test Environment:' + "\n"
        descriptionContent = descriptionContent + 'Workbanch code inspection environment.' + "\n"+ "\n"
        descriptionContent = descriptionContent + 'Assumptions and Constraints:' + "\n"
        descriptionContent = descriptionContent + 'Workbanch code inspection environment loaded with the latest code.'+ "\n"+ "\n"
        descriptionContent = descriptionContent + 'LLR Coverage Rationale:' + "\n"
        descriptionContent = descriptionContent + 'The LLRs are verified explicitly by code inspection and therefore covered by the test case.' + "\n" + "\n"
        descriptionContent = descriptionContent + 'Robustness/Normal test:' + "\n"
        descriptionContent = descriptionContent + 'Normal.' + "\n"
        return descriptionContent
'''--------------------------------------------------------------------'''


'''The function inputValidation() read from user the inputs paths'''
fieldValue = inputValidation()
 
'''Path of excel file that contains the LLR, function name and LLR name'''
ExeclAnalysisPath = fieldValue[0]

'''Save file name in variable myFile'''
myFile = ExeclAnalysisPath.rsplit('\\', 1)[1]

'execl PATH'
wb = load_workbook(ExeclAnalysisPath)

''' create new sheet for results '''
'''Check if sheet exist in dile - if exist do not create a new one'''
if not 'Ready_Analysis' in wb.sheetnames:
       wb.create_sheet('Ready_Analysis')

''' load the "Query1" sheet'''
sheet = wb.get_sheet_by_name('Query1')

'''Step number'''
stepCounter = 0

'''For loop that run until the cell in next row is NULL'''
for cell in sheet['A']:
    '''Check if the cell not beginning with 'RQ'(header)'''
    if not(str(cell.value).startswith("RQ")):
        stepCounter = stepCounter + 1 
        ReqContent = str(cell.value)
        
        '''Remove the last word in string, usually the last word is REQ000000'''
        ReqContent = ReqContent.rsplit(' ', 1)[0]
        
        '''----------s/es/ies verbs---------------------------------------------------------------'''
        
        '''Save in req the right split after "shall"'''
        req = ReqContent.rsplit('shall', 1)[1]
        
        string = req.split(' ')
        '''Save in str1 the verb that we want to edit'''
        str1 = string[1]
        '''Convert the verb to list and save it in charArray'''
        charArray = list(str1)

        '''Save the 4 last characters in char1-char4'''
        lengthStr = len(charArray)
        char1 = charArray[lengthStr-1]
        char2 = charArray[lengthStr-2]
        char3 = charArray[lengthStr-3]
        char4 = charArray[lengthStr-4]
        
        '''Indicate if the verb edited or not'''
        flag = 0
        '''if the last char is 'y' and before that one of the vowels'''
        if (char1 == 'y' and (char2 != 'a') and (char2 != 'u') and (char2 != 'e') and (char2 != 'i') and (char2 != 'o')):
             charArray = charArray[:-1]
             charArray.append('i')
             charArray.append('e')
             charArray.append('s')
             flag = 1
        
        elif ((char1 == 's') or (char1 == 'z') or (char2 == 'c' and char1 == 'h') or (char2 == 's' and char1 == 'h') or (char1 == 'x') or (char1 == 'o')):
             charArray.append('e')
             charArray.append('s')
             flag = 1

        elif (char4 == 'h' and char3 == 'a' and char2 == 'v' and char1 == 'e'):
             charArray = charArray[:-2]
             charArray.append('s')
             flag = 1

        elif (flag == 0):
            charArray.append('s')

        newString = ''.join(charArray)

        ReqContent = ReqContent.split('shall ')

        ReqContent[1].split(' ', 1)

        ReqContent[1] = ReqContent[1].split(' ', 1)[1]


        ReqContent = ReqContent[0]+ newString + ' ' + ReqContent[1]

        '''----------s/es/ies verbs---------------------------------------------------------------'''

        currentCol = cell.column
        currentRow = cell.row

        ReqName = sheet[chr(ord(currentCol)+1) + str(currentRow)].value
        
        '''Print for debug'''
       
        print ("the Req name is---> " + ReqName+ "\n")

        splitReqName = ReqName.split('::')
        fileName = splitReqName[0]
        functionName = ReqName.rsplit(' ', 1)[0]

        
        '''Print for debug'''
        print ("the fileName is---> " + fileName+ "\n")
        print ("the functionName is---> " + functionName+ "\n")

        ''' Buliding the final string '''
        finalString = ("open the file: " + fileName + ".cpp/h" + "\n")
        finalString = finalString + "Go to the method: " + functionName + "\n"
        finalString = finalString + "Verify that " + ReqContent
        '''Print for debug'''
        print (finalString+ "\n")

        ''' load the "Ready_Analysis" sheet'''
        Ready_Analysis_sheet = wb.get_sheet_by_name('Ready_Analysis')
        
        '''Function that initalize the description'''
        descriptionContent = descriptionInitialize(fileName)
        
        ''' update Ready_Analysis_sheet '''
        Ready_Analysis_sheet[str( chr(ord(currentCol))) + str(currentRow-1)].value = fieldValue[1]
        Ready_Analysis_sheet[str( chr(ord(currentCol)+1 )) + str(currentRow-1)].value = fileName
        Ready_Analysis_sheet[str( chr(ord(currentCol)+2 )) + str(currentRow-1)].value = "MANUAL"
        Ready_Analysis_sheet[str( chr(ord(currentCol)+3 )) + str(currentRow-1)].value = "Analysis"
        Ready_Analysis_sheet[str( chr(ord(currentCol)+4 )) + str(currentRow-1)].value = "Step " + str(stepCounter)
        Ready_Analysis_sheet[str( chr(ord(currentCol)+5 )) + str(currentRow-1)].value = finalString
        Ready_Analysis_sheet[str( chr(ord(currentCol)+6 )) + str(currentRow-1)].value = ReqContent
        Ready_Analysis_sheet[str( chr(ord(currentCol)+7 )) + str(currentRow-1)].value = descriptionContent
        

wb.save(myFile)
